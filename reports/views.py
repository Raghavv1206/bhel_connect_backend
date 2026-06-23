import io
from decimal import Decimal
from django.conf import settings
from django.db.models import Count, Sum, Q
from django.db.models.functions import Coalesce
from django.http import HttpResponse, Http404
from django.shortcuts import get_object_or_404
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

# Models
from smartbuy.models import Campaign, CampaignRegistration
from marketplace.models import Category, MarketplaceListing
from users.models import Employee

# Permissions
from users.permissions import IsAdminEmployee

# Report utilities
from reports.report_utils import (
    excel_add_bhel_header,
    excel_write_header_row,
    excel_write_data_rows,
    excel_write_summary_row,
    excel_auto_size_columns,
    pdf_build_header_elements,
    pdf_build_table,
    pdf_build_summary_paragraph,
    format_inr,
    format_report_date,
)

# openpyxl
from openpyxl import Workbook

# ReportLab
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer


class CampaignBuyersReportView(APIView):
    """
    API view to export a report of buyers in a campaign.
    Supports Excel (.xlsx) and PDF (.pdf) formats.
    Requires admin privileges.
    """
    permission_classes = [IsAuthenticated, IsAdminEmployee]

    def perform_content_negotiation(self, request, force=False):
        # Bypass DRF format content negotiation to avoid 404 for format=excel or format=pdf
        renderers = self.get_renderers()
        return (renderers[0], renderers[0].media_type)

    def get(self, request, campaign_id):
        # 1. Fetch Campaign
        campaign = get_object_or_404(Campaign, id=campaign_id)

        # 2. Parse format
        fmt = request.GET.get('format', 'excel').lower()
        if fmt not in ['excel', 'pdf']:
            return Response({"detail": "Invalid format. Must be 'excel' or 'pdf'."}, status=400)

        try:
            # 3. Retrieve registrations (buyers only: is_waitlisted=False, exclude cancelled)
            registrations = CampaignRegistration.objects.filter(
                campaign=campaign,
                is_waitlisted=False
            ).exclude(
                payment_status='cancelled'
            ).select_related('employee').order_by('reservation_date')

            # Calculate prices
            final_price = campaign.get_current_price()
            
            # Prepare data
            headers = [
                "Employee ID", "Name", "Department", "Email", 
                "Mobile", "Token Paid", "Final Price", "Remaining Due", "Reservation Date"
            ]

            raw_data = []
            for reg in registrations:
                emp = reg.employee
                token_paid = reg.token_amount or Decimal("0.00")
                
                if final_price is not None:
                    rem_due = max(Decimal("0.00"), final_price - token_paid)
                    final_price_val = final_price
                else:
                    rem_due = "N/A"
                    final_price_val = "N/A"

                raw_data.append([
                    emp.employee_id,
                    emp.name,
                    emp.department,
                    emp.email,
                    emp.mobile or "N/A",
                    token_paid,
                    final_price_val,
                    rem_due,
                    reg.reservation_date
                ])

            # 4. Generate Response based on format
            if fmt == 'excel':
                wb = Workbook()
                ws = wb.active
                ws.title = "Campaign Buyers"

                # Setup layout
                row_num = excel_add_bhel_header(ws, f"Buyers Report - {campaign.title}", len(headers))
                row_num = excel_write_header_row(ws, row_num, headers)

                # Process dates and format for Excel data rows
                excel_rows = []
                for r in raw_data:
                    row_copy = list(r)
                    # Convert reservation_date to string format
                    row_copy[8] = format_report_date(row_copy[8])
                    # If currency is N/A, keep it as string. Otherwise, convert Dec to float for openpyxl
                    if isinstance(row_copy[5], Decimal):
                        row_copy[5] = float(row_copy[5])
                    if isinstance(row_copy[6], Decimal):
                        row_copy[6] = float(row_copy[6])
                    if isinstance(row_copy[7], Decimal):
                        row_copy[7] = float(row_copy[7])
                    excel_rows.append(row_copy)

                row_num = excel_write_data_rows(ws, row_num, excel_rows, currency_columns=[5, 6, 7])
                excel_write_summary_row(ws, row_num, f"Total Buyers: {len(excel_rows)} employees", len(headers))
                excel_auto_size_columns(ws)

                buffer = io.BytesIO()
                wb.save(buffer)
                
                filename = f"campaign_{campaign_id}_buyers.xlsx"
                response = HttpResponse(
                    buffer.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response

            else:  # PDF format
                buffer = io.BytesIO()
                doc = SimpleDocTemplate(
                    buffer,
                    pagesize=landscape(A4),
                    rightMargin=1.5 * cm,
                    leftMargin=1.5 * cm,
                    topMargin=1.5 * cm,
                    bottomMargin=1.5 * cm
                )

                styles = getSampleStyleSheet()
                cell_style = ParagraphStyle(
                    'CellS',
                    parent=styles['Normal'],
                    fontSize=8,
                    leading=10,
                )
                cell_style_center = ParagraphStyle(
                    'CellSCenter',
                    parent=cell_style,
                    alignment=1,  # Center
                )
                cell_style_right = ParagraphStyle(
                    'CellSRight',
                    parent=cell_style,
                    alignment=2,  # Right
                )
                header_style = ParagraphStyle(
                    'HeaderS',
                    parent=styles['Normal'],
                    fontSize=9,
                    leading=11,
                    textColor=colors.white,
                    fontName='Helvetica-Bold',
                    alignment=1,  # Center
                )

                # Process data to wrap in Paragraphs
                formatted_headers = [Paragraph(h, header_style) for h in headers]
                formatted_rows = []

                for r in raw_data:
                    row_cells = []
                    # Col indices:
                    # 0: Emp ID, 1: Name, 2: Dept, 3: Email, 4: Mobile, 5: Token Paid, 6: Final Price, 7: Remaining Due, 8: Res Date
                    for idx, val in enumerate(r):
                        if idx in [5, 6, 7]:  # Currency
                            val_str = format_inr(val) if isinstance(val, (Decimal, float)) else str(val)
                            row_cells.append(Paragraph(val_str, cell_style_right))
                        elif idx == 8:  # Date
                            val_str = format_report_date(val)
                            row_cells.append(Paragraph(val_str, cell_style_center))
                        elif idx in [0, 4]:  # ID and Mobile
                            row_cells.append(Paragraph(str(val), cell_style_center))
                        else:  # Name, Dept, Email
                            row_cells.append(Paragraph(str(val), cell_style))
                    formatted_rows.append(row_cells)

                # Total printable width on landscape A4 is 26.7 cm
                # Column widths allocation:
                col_widths = [2.5*cm, 3.5*cm, 3.5*cm, 4.5*cm, 2.7*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm]

                elements = pdf_build_header_elements(f"Buyers Report - {campaign.title}", len(raw_data))
                
                table = pdf_build_table(formatted_headers, formatted_rows, col_widths=col_widths)
                elements.append(table)

                elements.append(Spacer(1, 4 * cm))
                summary_para = pdf_build_summary_paragraph(f"Total Buyers: {len(raw_data)} employees")
                elements.append(summary_para)

                doc.build(elements)
                
                filename = f"campaign_{campaign_id}_buyers.pdf"
                response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response

        except Exception as e:
            return Response({"detail": f"An error occurred while generating the report: {str(e)}"}, status=500)


class CampaignWaitlistReportView(APIView):
    """
    API view to export a report of waitlisted employees in a campaign.
    Supports Excel (.xlsx) and PDF (.pdf) formats.
    Requires admin privileges.
    """
    permission_classes = [IsAuthenticated, IsAdminEmployee]

    def perform_content_negotiation(self, request, force=False):
        # Bypass DRF format content negotiation to avoid 404 for format=excel or format=pdf
        renderers = self.get_renderers()
        return (renderers[0], renderers[0].media_type)

    def get(self, request, campaign_id):
        # 1. Fetch Campaign
        campaign = get_object_or_404(Campaign, id=campaign_id)

        # 2. Parse format
        fmt = request.GET.get('format', 'excel').lower()
        if fmt not in ['excel', 'pdf']:
            return Response({"detail": "Invalid format. Must be 'excel' or 'pdf'."}, status=400)

        try:
            # 3. Retrieve waitlisted registrations (is_waitlisted=True, exclude cancelled)
            registrations = CampaignRegistration.objects.filter(
                campaign=campaign,
                is_waitlisted=True
            ).exclude(
                payment_status='cancelled'
            ).select_related('employee').order_by('waitlist_position', 'reservation_date')

            # Prepare data
            headers = [
                "Employee ID", "Name", "Department", "Email", "Mobile", "Waitlist Position", "Join Date"
            ]

            raw_data = []
            for reg in registrations:
                emp = reg.employee
                raw_data.append([
                    emp.employee_id,
                    emp.name,
                    emp.department,
                    emp.email,
                    emp.mobile or "N/A",
                    reg.waitlist_position or "N/A",
                    reg.reservation_date
                ])

            # 4. Generate Response based on format
            if fmt == 'excel':
                wb = Workbook()
                ws = wb.active
                ws.title = "Campaign Waitlist"

                # Setup layout
                row_num = excel_add_bhel_header(ws, f"Waitlist Report - {campaign.title}", len(headers))
                row_num = excel_write_header_row(ws, row_num, headers)

                # Process dates for Excel data rows
                excel_rows = []
                for r in raw_data:
                    row_copy = list(r)
                    row_copy[6] = format_report_date(row_copy[6])
                    excel_rows.append(row_copy)

                row_num = excel_write_data_rows(ws, row_num, excel_rows)
                excel_write_summary_row(ws, row_num, f"Total Waitlisted: {len(excel_rows)} employees", len(headers))
                excel_auto_size_columns(ws)

                buffer = io.BytesIO()
                wb.save(buffer)

                filename = f"campaign_{campaign_id}_waitlist.xlsx"
                response = HttpResponse(
                    buffer.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response

            else:  # PDF format (Portrait since it has 7 columns)
                buffer = io.BytesIO()
                doc = SimpleDocTemplate(
                    buffer,
                    pagesize=A4,
                    rightMargin=1.5 * cm,
                    leftMargin=1.5 * cm,
                    topMargin=1.5 * cm,
                    bottomMargin=1.5 * cm
                )

                styles = getSampleStyleSheet()
                cell_style = ParagraphStyle(
                    'CellSWait',
                    parent=styles['Normal'],
                    fontSize=8,
                    leading=10,
                )
                cell_style_center = ParagraphStyle(
                    'CellSWaitCenter',
                    parent=cell_style,
                    alignment=1,  # Center
                )
                header_style = ParagraphStyle(
                    'HeaderSWait',
                    parent=styles['Normal'],
                    fontSize=9,
                    leading=11,
                    textColor=colors.white,
                    fontName='Helvetica-Bold',
                    alignment=1,  # Center
                )

                # Process data to wrap in Paragraphs
                formatted_headers = [Paragraph(h, header_style) for h in headers]
                formatted_rows = []

                for r in raw_data:
                    row_cells = []
                    # Col indices:
                    # 0: Emp ID, 1: Name, 2: Dept, 3: Email, 4: Mobile, 5: Waitlist Position, 6: Join Date
                    for idx, val in enumerate(r):
                        if idx == 6:  # Date
                            val_str = format_report_date(val)
                            row_cells.append(Paragraph(val_str, cell_style_center))
                        elif idx in [0, 4, 5]:  # ID, Mobile, Position
                            row_cells.append(Paragraph(str(val), cell_style_center))
                        else:  # Name, Dept, Email
                            row_cells.append(Paragraph(str(val), cell_style))
                    formatted_rows.append(row_cells)

                # Total printable width on portrait A4 is 18.0 cm
                # Column widths allocation:
                col_widths = [2.5*cm, 3.0*cm, 3.0*cm, 3.5*cm, 2.5*cm, 1.5*cm, 2.0*cm]

                elements = pdf_build_header_elements(f"Waitlist Report - {campaign.title}", len(raw_data))
                
                table = pdf_build_table(formatted_headers, formatted_rows, col_widths=col_widths)
                elements.append(table)

                elements.append(Spacer(1, 4 * cm))
                summary_para = pdf_build_summary_paragraph(f"Total Waitlisted: {len(raw_data)} employees")
                elements.append(summary_para)

                doc.build(elements)

                filename = f"campaign_{campaign_id}_waitlist.pdf"
                response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response

        except Exception as e:
            return Response({"detail": f"An error occurred while generating the report: {str(e)}"}, status=500)


class MarketplaceSummaryReportView(APIView):
    """
    API view to export a comprehensive Marketplace Summary Excel workbook.
    Contains three sheets:
    1. Summary — Overview statistics of listings
    2. Category Breakdown — Listings count and sold value per category
    3. Top Sellers — Listing and sales activity per employee
    Requires admin privileges.
    """
    permission_classes = [IsAuthenticated, IsAdminEmployee]

    def perform_content_negotiation(self, request, force=False):
        # Bypass DRF format content negotiation to avoid 404 for format=excel
        renderers = self.get_renderers()
        return (renderers[0], renderers[0].media_type)

    def get(self, request):
        fmt = request.GET.get('format', 'excel').lower()
        if fmt != 'excel':
            return Response({"detail": "Invalid format. Marketplace summary is only available in 'excel' format."}, status=400)

        try:
            # 1. Gather Sheet 1: Summary counts
            counts = MarketplaceListing.objects.aggregate(
                total=Count('id'),
                active=Count('id', filter=Q(status__in=['approved', 'available', 'reserved'])),
                sold=Count('id', filter=Q(status='sold')),
                pending=Count('id', filter=Q(status='pending')),
                rejected=Count('id', filter=Q(status='rejected'))
            )

            # 2. Gather Sheet 2: Category Breakdown
            categories_data = Category.objects.annotate(
                listing_count=Count('listings'),
                sold_count=Count('listings', filter=Q(listings__status='sold')),
                total_sold_value=Coalesce(Sum('listings__price', filter=Q(listings__status='sold')), Decimal('0.00'))
            ).order_by('display_order', 'name')

            # 3. Gather Sheet 3: Top Sellers (Sellers with at least 1 listing)
            sellers_data = Employee.objects.annotate(
                active_listings=Count('marketplace_listings', filter=Q(marketplace_listings__status__in=['approved', 'available', 'reserved'])),
                sold_listings=Count('marketplace_listings', filter=Q(marketplace_listings__status='sold'))
            ).filter(
                Q(active_listings__gt=0) | Q(sold_listings__gt=0)
            ).order_by('-sold_listings', '-active_listings')

            # Create workbook
            wb = Workbook()

            # ──── Sheet 1: Summary ──────────────────────────────────────────
            ws_summary = wb.active
            ws_summary.title = "Summary"
            
            summary_headers = ["Metric", "Count"]
            row_num = excel_add_bhel_header(ws_summary, "Marketplace Listings Summary", len(summary_headers))
            row_num = excel_write_header_row(ws_summary, row_num, summary_headers)
            
            summary_rows = [
                ["Total Listings", counts['total']],
                ["Active Listings (Approved, Available, Reserved)", counts['active']],
                ["Sold Listings", counts['sold']],
                ["Pending Review Listings", counts['pending']],
                ["Rejected Listings", counts['rejected']]
            ]
            row_num = excel_write_data_rows(ws_summary, row_num, summary_rows)
            excel_auto_size_columns(ws_summary)

            # ──── Sheet 2: Category Breakdown ───────────────────────────────
            ws_categories = wb.create_sheet(title="Category Breakdown")
            cat_headers = ["Category Name", "Listing Count", "Sold Count", "Total Sold Value"]
            row_num = excel_add_bhel_header(ws_categories, "Category Sales Breakdown", len(cat_headers))
            row_num = excel_write_header_row(ws_categories, row_num, cat_headers)

            cat_rows = []
            for cat in categories_data:
                # Convert Dec to float for openpyxl
                sold_val = float(cat.total_sold_value) if isinstance(cat.total_sold_value, Decimal) else cat.total_sold_value
                cat_rows.append([
                    str(cat),  # Use the standard __str__ which has parent hierarchy
                    cat.listing_count,
                    cat.sold_count,
                    sold_val
                ])
            row_num = excel_write_data_rows(ws_categories, row_num, cat_rows, currency_columns=[3])
            excel_auto_size_columns(ws_categories)

            # ──── Sheet 3: Top Sellers ──────────────────────────────────────
            ws_sellers = wb.create_sheet(title="Top Sellers")
            seller_headers = ["Employee Name", "Department", "Active Listings", "Sold Listings"]
            row_num = excel_add_bhel_header(ws_sellers, "Top Sellers Leaderboard", len(seller_headers))
            row_num = excel_write_header_row(ws_sellers, row_num, seller_headers)

            seller_rows = []
            for seller in sellers_data:
                seller_rows.append([
                    seller.name,
                    seller.department,
                    seller.active_listings,
                    seller.sold_listings
                ])
            row_num = excel_write_data_rows(ws_sellers, row_num, seller_rows)
            excel_auto_size_columns(ws_sellers)

            # Write workbook to response
            buffer = io.BytesIO()
            wb.save(buffer)

            filename = "marketplace_summary.xlsx"
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        except Exception as e:
            return Response({"detail": f"An error occurred while generating the report: {str(e)}"}, status=500)
