"""
reports/report_utils.py
Shared Excel (openpyxl) and PDF (ReportLab) formatting utilities used by all
report generator views. Centralises branding, colours, and layout logic so
every exported document looks consistent.

BHEL branding colours:
  - Primary Blue:  #003366
  - Header Row BG: #003366 with white text
  - Alternating PDF rows: #F2F6FA (light blue-grey)
"""

import os
from decimal import Decimal
from datetime import datetime

from django.conf import settings
from django.utils import timezone

# ── openpyxl imports ────────────────────────────────────────────────────────
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── ReportLab imports ───────────────────────────────────────────────────────
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm, cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT


# ═══════════════════════════════════════════════════════════════════════════
# COLOUR / STYLE CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════

# Excel colours
BHEL_BLUE_HEX = "003366"
HEADER_FILL = PatternFill(start_color=BHEL_BLUE_HEX, end_color=BHEL_BLUE_HEX, fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, color=BHEL_BLUE_HEX, size=14)
SUMMARY_FONT = Font(name="Calibri", bold=True, color=BHEL_BLUE_HEX, size=11)
DATA_FONT = Font(name="Calibri", size=10)
CURRENCY_FORMAT = '#,##0.00'
DATE_FORMAT = 'DD-MMM-YYYY'
THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)
ALT_ROW_FILL = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")

# PDF colours
BHEL_BLUE_RGB = colors.HexColor("#003366")
PDF_HEADER_BG = colors.HexColor("#003366")
PDF_HEADER_TEXT = colors.white
PDF_ALT_ROW = colors.HexColor("#F2F6FA")
PDF_GRID_COLOR = colors.HexColor("#CCCCCC")


# ═══════════════════════════════════════════════════════════════════════════
# EXCEL HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════

def excel_add_bhel_header(ws, title_text, column_count):
    """
    Add a BHEL branded header block at the top of an Excel worksheet.
    Row 1: "BHEL CONNECT MARKETPLACE" in large bold blue
    Row 2: Report title
    Row 3: Generation timestamp
    Row 4: Empty separator
    Returns the next usable row number (5).
    """
    # Row 1 — Organisation name
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
    cell = ws.cell(row=1, column=1, value="BHEL CONNECT MARKETPLACE")
    cell.font = Font(name="Calibri", bold=True, color=BHEL_BLUE_HEX, size=16)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    # Row 2 — Report title
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=column_count)
    cell = ws.cell(row=2, column=1, value=title_text)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")

    # Row 3 — Timestamp
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=column_count)
    timestamp_str = timezone.localtime(timezone.now()).strftime("%d-%b-%Y %I:%M %p IST")
    cell = ws.cell(row=3, column=1, value=f"Generated: {timestamp_str}")
    cell.font = Font(name="Calibri", italic=True, color="666666", size=9)
    cell.alignment = Alignment(horizontal="left")

    # Row 4 — separator (empty)
    ws.row_dimensions[4].height = 6

    return 5  # Next available row


def excel_write_header_row(ws, row_num, headers):
    """
    Write a styled header row with BHEL blue background and white bold text.
    Freezes the pane below the header row for easy scrolling.
    """
    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=header_text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Set row height for header
    ws.row_dimensions[row_num].height = 28

    # Freeze pane below header row
    ws.freeze_panes = ws.cell(row=row_num + 1, column=1)

    return row_num + 1  # Next data row


def excel_write_data_rows(ws, start_row, data_rows, currency_columns=None):
    """
    Write data rows with alternating row shading, thin borders,
    and optional currency formatting on specified columns (0-indexed).
    Returns the next available row number after the last data row.
    """
    currency_columns = currency_columns or []
    current_row = start_row

    for row_idx, row_data in enumerate(data_rows):
        for col_idx, value in enumerate(row_data):
            cell = ws.cell(row=current_row, column=col_idx + 1, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

            # Apply currency format to designated columns
            if col_idx in currency_columns:
                cell.number_format = CURRENCY_FORMAT
                cell.alignment = Alignment(horizontal="right", vertical="center")

            # Alternating row fill
            if row_idx % 2 == 1:
                cell.fill = ALT_ROW_FILL

        ws.row_dimensions[current_row].height = 22
        current_row += 1

    return current_row


def excel_write_summary_row(ws, row_num, summary_text, column_count):
    """
    Write a bold summary/footer row spanning all columns.
    """
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=column_count)
    cell = ws.cell(row=row_num, column=1, value=summary_text)
    cell.font = SUMMARY_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row_num].height = 26
    return row_num + 1


def excel_auto_size_columns(ws, min_width=10, max_width=40):
    """
    Auto-size worksheet columns based on content length.
    Clamps between min_width and max_width to prevent extreme widths.
    """
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                max_length = max(max_length, cell_len)
        # Clamp the adjusted width
        adjusted = min(max(max_length + 3, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted


# ═══════════════════════════════════════════════════════════════════════════
# PDF HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════

def _get_bhel_logo_path():
    """
    Resolve the path to the BHEL logo image for PDF headers.
    Checks both SVG and PNG candidates in STATIC_ROOT and project-level static directory.
    Returns None if no matching file exists.
    """
    candidates = [
        'BHEL_logo.svg',
        'bhel_logo.svg',
        'BHEL_logo.png',
        'bhel_logo.png'
    ]

    # Try STATIC_ROOT first (production after collectstatic)
    if settings.STATIC_ROOT:
        for candidate in candidates:
            path = os.path.join(str(settings.STATIC_ROOT), 'images', candidate)
            if os.path.isfile(path):
                return path

    # Fallback: project-level static directory (development)
    for candidate in candidates:
        path = os.path.join(str(settings.BASE_DIR), 'static', 'images', candidate)
        if os.path.isfile(path):
            return path

    return None


def pdf_build_header_elements(title_text, record_count=None):
    """
    Build a list of ReportLab flowable elements for the PDF header:
    - BHEL logo (or text fallback)
    - Report title
    - Generation timestamp
    - Record count (optional)
    Returns a list of flowable elements to prepend to the document.
    """
    styles = getSampleStyleSheet()
    elements: list = []

    # Define custom styles for the header
    brand_style = ParagraphStyle(
        'BHELBrand',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=BHEL_BLUE_RGB,
        spaceAfter=2 * mm,
        leading=22,
    )
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=BHEL_BLUE_RGB,
        spaceAfter=2 * mm,
        leading=18,
    )
    meta_style = ParagraphStyle(
        'ReportMeta',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor("#666666"),
        spaceAfter=4 * mm,
    )

    # Attempt to include the BHEL logo image
    logo_path = _get_bhel_logo_path()
    if logo_path:
        try:
            if logo_path.lower().endswith('.svg'):
                # Render SVG drawing using svglib
                from svglib.svglib import svg2rlg
                logo = svg2rlg(logo_path)
                if logo:
                    logo.hAlign = 'LEFT'
                    # Scale to fit ~40mm width and ~12mm height
                    factor_x = (40 * mm) / logo.width if logo.width > 0 else 1.0
                    factor_y = (12 * mm) / logo.height if logo.height > 0 else 1.0
                    factor = min(factor_x, factor_y)
                    logo.width = logo.width * factor
                    logo.height = logo.height * factor
                    logo.scale(factor, factor)
                    elements.append(logo)
                    elements.append(Spacer(1, 3 * mm))
                else:
                    elements.append(Paragraph("BHEL CONNECT MARKETPLACE", brand_style))
            else:
                # Regular raster image loading (PNG/JPG)
                logo = Image(logo_path, width=40 * mm, height=12 * mm)
                logo.hAlign = 'LEFT'
                elements.append(logo)
                elements.append(Spacer(1, 3 * mm))
        except Exception:
            # If image loading fails, fall back to text header
            elements.append(Paragraph("BHEL CONNECT MARKETPLACE", brand_style))
    else:
        # Text fallback when no logo file is available
        elements.append(Paragraph("BHEL CONNECT MARKETPLACE", brand_style))

    # Report title
    elements.append(Paragraph(title_text, title_style))

    # Generation timestamp and optional record count
    timestamp_str = timezone.localtime(timezone.now()).strftime("%d-%b-%Y %I:%M %p IST")
    meta_parts = [f"Generated: {timestamp_str}"]
    if record_count is not None:
        meta_parts.append(f"Total Records: {record_count}")
    elements.append(Paragraph(" | ".join(meta_parts), meta_style))

    # Horizontal rule via a thin table
    hr_data = [['', '']]
    hr_table = Table(hr_data, colWidths=['*', '*'])
    hr_table.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, 0), 1, BHEL_BLUE_RGB),
    ]))
    elements.append(hr_table)
    elements.append(Spacer(1, 4 * mm))

    return elements


def pdf_build_table(headers, data_rows, col_widths=None):
    """
    Build a styled ReportLab Table with:
    - Blue header row with white text
    - Alternating row colours
    - Grid lines
    Returns the Table flowable.
    """
    # Combine headers and data into a single table dataset
    table_data = [headers] + data_rows

    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Build the style commands
    style_commands = [
        # Header row styling
        ('BACKGROUND', (0, 0), (-1, 0), PDF_HEADER_BG),
        ('TEXTCOLOR', (0, 0), (-1, 0), PDF_HEADER_TEXT),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),

        # Data row styling
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, PDF_GRID_COLOR),

        # Outer border
        ('BOX', (0, 0), (-1, -1), 1, BHEL_BLUE_RGB),
    ]

    # Alternating row colours for data rows
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            style_commands.append(
                ('BACKGROUND', (0, i), (-1, i), PDF_ALT_ROW)
            )

    table.setStyle(TableStyle(style_commands))
    return table


def pdf_build_summary_paragraph(text):
    """
    Build a styled summary paragraph for PDF footers.
    """
    styles = getSampleStyleSheet()
    summary_style = ParagraphStyle(
        'SummaryStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=BHEL_BLUE_RGB,
        fontName='Helvetica-Bold',
        spaceBefore=4 * mm,
        spaceAfter=2 * mm,
    )
    return Paragraph(text, summary_style)


# ═══════════════════════════════════════════════════════════════════════════
# SHARED DATA FORMATTING HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def format_inr(amount):
    """Format a Decimal/float as an INR string with commas: ₹1,23,456.00"""
    if amount is None:
        return "₹0.00"
    try:
        value = float(amount)
        # Indian number format: last 3 digits, then groups of 2
        # Use a simple approach: format with 2 decimal places
        formatted = f"{value:,.2f}"
        return f"₹{formatted}"
    except (ValueError, TypeError):
        return "₹0.00"


def format_report_date(dt):
    """Format a datetime object for display in reports: '07-Jun-2026'."""
    if dt is None:
        return "N/A"
    try:
        local_dt = timezone.localtime(dt)
        return local_dt.strftime("%d-%b-%Y")
    except Exception:
        return str(dt)


def generate_campaign_bookings_excel(campaign):
    """
    Generates an Excel workbook with two sheets:
    1. Confirmed Bookings (Approved payment, not waitlisted)
    2. Pending & Waitlisted (Pending, rejected, or cancelled payments, or waitlisted)
    Returns a bytes object of the generated workbook.
    """
    import io
    from decimal import Decimal
    from openpyxl import Workbook
    from smartbuy.models import CampaignRegistration
    
    wb = Workbook()
    
    # ── Sheet 1: Confirmed Bookings ──────────────────────────────────────────
    ws_confirmed = wb.active
    ws_confirmed.title = "Confirmed Bookings"
    
    confirmed_headers = [
        "Employee ID", "Name", "Department", "Email", 
        "Mobile", "Token Paid (₹)", "Final Price (₹)", "Remaining Due (₹)", "Booking Date"
    ]
    
    # Get confirmed bookings
    confirmed_regs = CampaignRegistration.objects.filter(
        campaign=campaign,
        payment_status='approved',
        is_waitlisted=False
    ).select_related('employee').order_by('reservation_date')
    
    final_price = campaign.get_current_price()
    
    row_num = excel_add_bhel_header(ws_confirmed, f"Confirmed Bookings - {campaign.title}", len(confirmed_headers))
    row_num = excel_write_header_row(ws_confirmed, row_num, confirmed_headers)
    
    confirmed_rows = []
    for reg in confirmed_regs:
        emp = reg.employee
        token_paid = reg.token_amount or Decimal("0.00")
        
        if final_price is not None:
            rem_due = max(Decimal("0.00"), final_price - token_paid)
            final_price_val = final_price
        else:
            rem_due = Decimal("0.00")
            final_price_val = Decimal("0.00")
            
        confirmed_rows.append([
            emp.employee_id,
            emp.name,
            emp.department,
            emp.email,
            emp.mobile or "N/A",
            float(token_paid),
            float(final_price_val),
            float(rem_due),
            format_report_date(reg.reservation_date)
        ])
        
    row_num = excel_write_data_rows(ws_confirmed, row_num, confirmed_rows, currency_columns=[5, 6, 7])
    excel_write_summary_row(ws_confirmed, row_num, f"Total Confirmed: {len(confirmed_rows)} employees", len(confirmed_headers))
    excel_auto_size_columns(ws_confirmed)
    
    # ── Sheet 2: Pending & Waitlisted ────────────────────────────────────────
    ws_pending = wb.create_sheet(title="Pending & Waitlisted")
    
    pending_headers = [
        "Employee ID", "Name", "Department", "Email", 
        "Mobile", "Status", "Waitlist Position", "Token Deposit (₹)", "Cashfree Order ID", "Slot Expiry Date", "Registration Date"
    ]
    
    # Get pending and waitlisted (exclude confirmed)
    pending_regs = CampaignRegistration.objects.filter(
        campaign=campaign
    ).exclude(
        payment_status='approved',
        is_waitlisted=False
    ).select_related('employee').order_by('is_waitlisted', 'waitlist_position', 'reservation_date')
    
    row_num = excel_add_bhel_header(ws_pending, f"Pending & Waitlisted Bookings - {campaign.title}", len(pending_headers))
    row_num = excel_write_header_row(ws_pending, row_num, pending_headers)
    
    pending_rows = []
    for reg in pending_regs:
        emp = reg.employee
        status_str = "Waitlisted" if reg.is_waitlisted else reg.get_payment_status_display()
        token_amt = reg.token_amount or Decimal("0.00")
        
        pending_rows.append([
            emp.employee_id,
            emp.name,
            emp.department,
            emp.email,
            emp.mobile or "N/A",
            status_str,
            reg.waitlist_position or "N/A",
            float(token_amt),
            reg.cashfree_order_id or "N/A",
            format_report_date(reg.slot_expiry_date) if reg.slot_expiry_date else "N/A",
            format_report_date(reg.reservation_date)
        ])
        
    row_num = excel_write_data_rows(ws_pending, row_num, pending_rows, currency_columns=[7])
    excel_write_summary_row(ws_pending, row_num, f"Total Pending & Waitlisted: {len(pending_rows)} registrations", len(pending_headers))
    excel_auto_size_columns(ws_pending)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
